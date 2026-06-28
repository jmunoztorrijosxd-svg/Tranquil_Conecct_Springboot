import os
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q
from .models import Post
from chat.models import Usuario

# --- LIBRERÍAS PARA EXCEL ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage  # Para insertar imágenes en Excel

# --- LIBRERÍAS PARA PDF ---
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

# 1. VISTA DEL PANEL DE ADMINISTRACIÓN (LISTADO Y BÚSQUEDA)
def admin_posts_list(request):
    query = request.GET.get('q', '').strip() 
    posts = Post.objects.all()

    if query:
        # Buscamos por contenido o por ID de usuario si es un número
        if query.isdigit():
            posts = posts.filter(Q(contenido__icontains=query) | Q(usuario_id=query))
        else:
            posts = posts.filter(contenido__icontains=query)

    posts = posts.order_by('-fecha_publicacion')

    # Cargamos los nombres de usuario para mejorar la vista de moderación
    user_ids = set(posts.values_list('usuario_id', flat=True))
    usuarios = Usuario.objects.filter(id__in=user_ids)
    user_map = {u.id: u.nombre for u in usuarios}

    # Adjuntamos un atributo temporal `usuario_nombre` a cada objeto Post para usar en la plantilla
    for post in posts:
        post.usuario_nombre = user_map.get(post.usuario_id, '')

    return render(request, 'posts/admin_list.html', {'posts': posts})


# 2. ELIMINAR PUBLICACIÓN
def delete_post(request, pk):
    post = get_object_or_404(Post, pk=pk)
    post.delete()
    return redirect('admin_posts_list')


# 3. EXPORTAR EXCEL CON IMÁGENES Y FORMATO
def export_posts_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Tranquil Connect"
    
    # Encabezados (Columna E para la Imagen)
    headers = ['ID', 'Contenido', 'ID Usuario', 'Fecha', 'Imagen']
    ws.append(headers)
    
    # Estilo visual para el encabezado (Azul)
    blue_fill = PatternFill(start_color="A0C4FF", end_color="A0C4FF", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center")

    fila_actual = 2
    for post in Post.objects.all():
        fecha = post.fecha_publicacion.replace(tzinfo=None).strftime("%d/%m/%Y %H:%M") if post.fecha_publicacion else "Sin fecha"
        ws.append([post.id, post.contenido, post.usuario_id, fecha])
        
        # --- INSERTAR IMAGEN EN EXCEL ---
        if post.imagen:
            try:
                ruta_imagen = os.path.join(settings.MEDIA_ROOT, str(post.imagen))
                if os.path.exists(ruta_imagen):
                    img = ExcelImage(ruta_imagen)
                    # Redimensionamos la imagen para que quepa en la celda
                    img.width = 80
                    img.height = 60
                    
                    # Ajustamos la altura de la fila para que la imagen no se vea pequeña
                    ws.row_dimensions[fila_actual].height = 50 
                    # Colocamos la imagen en la celda E
                    ws.add_image(img, f'E{fila_actual}')
            except:
                ws.cell(row=fila_actual, column=5).value = "[Error de imagen]"
        
        fila_actual += 1

    # Ajuste manual de anchos de columna
    ws.column_dimensions['B'].width = 40  # Contenido
    ws.column_dimensions['D'].width = 20  # Fecha
    ws.column_dimensions['E'].width = 15  # Imagen

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_publicaciones.xlsx"'
    wb.save(response)
    return response


def edit_post(request, pk):
    post = get_object_or_404(Post, pk=pk)
    
    if request.method == "POST":
        # Obtenemos los datos del formulario
        post.contenido = request.POST.get('contenido')
        
        # Si el usuario sube una nueva imagen, la actualizamos
        if request.FILES.get('imagen'):
            post.imagen = request.FILES.get('imagen')
            
        post.save()
        return redirect('admin_posts_list')

    return render(request, 'posts/edit_post.html', {'post': post})


# 4. EXPORTAR PDF CON IMÁGENES Y FORMATO
def export_posts_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_publicaciones.pdf"'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    # Cabecera del Documento
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 50, "Reporte de Moderación - Tranquil Connect")
    
    y = height - 100
    
    for post in Post.objects.all():
        # Información del Post
        p.setFont("Helvetica-Bold", 10)
        p.drawString(100, y, f"ID: {post.id} | Usuario ID: {post.usuario_id}")
        
        p.setFont("Helvetica", 9)
        # Cortamos el contenido si es muy largo para que no se salga del PDF
        resumen = (post.contenido[:75] + '...') if len(post.contenido) > 75 else post.contenido
        p.drawString(100, y - 15, f"Contenido: {resumen}")

        # --- DIBUJAR IMAGEN EN EL PDF ---
        if post.imagen:
            try:
                ruta_imagen = os.path.join(settings.MEDIA_ROOT, str(post.imagen))
                if os.path.exists(ruta_imagen):
                    # p.drawImage(ruta, x, y, ancho, alto)
                    p.drawImage(ruta_imagen, 420, y - 35, width=80, height=60, preserveAspectRatio=True)
            except:
                p.drawString(420, y, "[Imagen no disponible]")

        # Línea divisoria entre posts
        p.setStrokeColor(colors.lightgrey)
        p.line(100, y - 45, 520, y - 45)
        
        y -= 85 # Bajamos la posición para el siguiente post
        
        # Si se acaba el espacio en la hoja, creamos una nueva
        if y < 100:
            p.showPage()
            y = height - 50

    p.showPage()
    p.save()
    return response